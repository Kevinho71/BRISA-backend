"""Servicio para lógica de negocio de Estudiantes"""

from sqlalchemy.orm import Session
from typing import Optional, List
from fastapi import HTTPException, UploadFile
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import pandas as pd

from app.modules.estudiantes.repositories.estudiante_repository import EstudianteRepository
from app.modules.estudiantes.dto.estudiante_dto import (
    EstudianteCreateDTO,
    EstudianteUpdateDTO,
    EstudianteResponseDTO,
    EstudianteListDTO,
    CursoBasicoDTO,
    CambiarEstadoDTO,
    CambiarEstadoResponseDTO,
    ImportResultDTO,
    EstadoEstudiante
)


class EstudianteService:
    """Servicio para gestión de estudiantes"""
    
    # ============= CRUD Básico =============
    
    @staticmethod
    def crear_estudiante(db: Session, datos: EstudianteCreateDTO) -> EstudianteResponseDTO:
        """RF-EST-001: Crear Estudiante"""
        datos_dict = datos.model_dump()
        
        # Convertir enum a string
        if 'estado' in datos_dict and isinstance(datos_dict['estado'], EstadoEstudiante):
            datos_dict['estado'] = datos_dict['estado'].value
        
        estudiante = EstudianteRepository.crear_estudiante(db, datos_dict)
        return EstudianteService._convertir_a_dto(estudiante)
    
    @staticmethod
    def obtener_todos(db: Session, skip: int = 0, limit: int = 100) -> EstudianteListDTO:
        """RF-EST-002: Listar Todos los Estudiantes"""
        estudiantes = EstudianteRepository.obtener_todos(db, skip, limit)
        total = EstudianteRepository.contar_total(db)
        
        estudiantes_dto = [EstudianteService._convertir_a_dto(est) for est in estudiantes]
        
        return EstudianteListDTO(
            total=total,
            estudiantes=estudiantes_dto
        )
    
    @staticmethod
    def obtener_por_id(db: Session, id_estudiante: int) -> EstudianteResponseDTO:
        """RF-EST-003: Obtener Estudiante por ID"""
        estudiante = EstudianteRepository.obtener_por_id(db, id_estudiante)
        
        if not estudiante:
            raise HTTPException(
                status_code=404,
                detail=f"Estudiante con ID {id_estudiante} no encontrado"
            )
        
        return EstudianteService._convertir_a_dto(estudiante)
    
    @staticmethod
    def actualizar_estudiante(
        db: Session,
        id_estudiante: int,
        datos: EstudianteUpdateDTO
    ) -> EstudianteResponseDTO:
        """RF-EST-004: Actualizar Estudiante"""
        # Verificar que exista
        estudiante_existe = EstudianteRepository.obtener_por_id(db, id_estudiante)
        if not estudiante_existe:
            raise HTTPException(
                status_code=404,
                detail=f"Estudiante con ID {id_estudiante} no encontrado"
            )
        
        # Preparar datos para actualizar (solo los proporcionados)
        datos_dict = datos.model_dump(exclude_unset=True)
        
        # Convertir enum a string
        if 'estado' in datos_dict and isinstance(datos_dict['estado'], EstadoEstudiante):
            datos_dict['estado'] = datos_dict['estado'].value
        
        estudiante_actualizado = EstudianteRepository.actualizar_estudiante(
            db, id_estudiante, datos_dict
        )
        
        return EstudianteService._convertir_a_dto(estudiante_actualizado)
    
    @staticmethod
    def eliminar_estudiante(db: Session, id_estudiante: int) -> dict:
        """RF-EST-005: Eliminar Estudiante"""
        # Verificar que exista
        estudiante = EstudianteRepository.obtener_por_id(db, id_estudiante)
        if not estudiante:
            raise HTTPException(
                status_code=404,
                detail=f"Estudiante con ID {id_estudiante} no encontrado"
            )
        
        nombre_completo = f"{estudiante.nombres} {estudiante.apellido_paterno}"
        
        # Eliminar (cascada automática en asignaciones)
        EstudianteRepository.eliminar_estudiante(db, id_estudiante)
        
        return {
            "mensaje": f"Estudiante {nombre_completo} eliminado exitosamente",
            "id_estudiante": id_estudiante
        }
    
    # ============= Gestión de Estados =============
    
    @staticmethod
    def cambiar_estado(
        db: Session,
        id_estudiante: int,
        datos: CambiarEstadoDTO
    ) -> CambiarEstadoResponseDTO:
        """RF-EST-006: Cambiar Estado de Estudiante"""
        resultado = EstudianteRepository.cambiar_estado(
            db, id_estudiante, datos.nuevo_estado.value
        )
        
        if not resultado:
            raise HTTPException(
                status_code=404,
                detail=f"Estudiante con ID {id_estudiante} no encontrado"
            )
        
        estudiante = resultado["estudiante"]
        estado_anterior = resultado["estado_anterior"]
        estado_nuevo = resultado["estado_nuevo"]
        
        if estado_anterior == estado_nuevo:
            raise HTTPException(
                status_code=400,
                detail=f"El estudiante ya tiene el estado {estado_nuevo}"
            )
        
        return CambiarEstadoResponseDTO(
            mensaje=f"Estado cambiado de {estado_anterior} a {estado_nuevo}",
            id_estudiante=id_estudiante,
            estado_anterior=estado_anterior,
            estado_nuevo=estado_nuevo,
            nombre_estudiante=f"{estudiante.nombres} {estudiante.apellido_paterno}"
        )
    
    @staticmethod
    def listar_por_estado(
        db: Session,
        estado: str,
        skip: int = 0,
        limit: int = 100
    ) -> EstudianteListDTO:
        """RF-EST-007: Listar Estudiantes por Estado"""
        # Validar estado
        estados_validos = [e.value for e in EstadoEstudiante]
        if estado not in estados_validos:
            raise HTTPException(
                status_code=400,
                detail=f"Estado '{estado}' no válido. Estados válidos: {', '.join(estados_validos)}"
            )
        
        estudiantes = EstudianteRepository.obtener_por_estado(db, estado, skip, limit)
        total = EstudianteRepository.contar_por_estado(db, estado)
        
        estudiantes_dto = [EstudianteService._convertir_a_dto(est) for est in estudiantes]
        
        return EstudianteListDTO(
            total=total,
            estudiantes=estudiantes_dto
        )
    
    # ============= Filtros por Gestión =============
    
    @staticmethod
    def listar_por_gestion(
        db: Session,
        gestion: Optional[str] = None,
        nivel: Optional[str] = None,
        id_curso: Optional[int] = None,
        skip: int = 0,
        limit: int = 100
    ) -> EstudianteListDTO:
        """RF-EST-008: Listar Estudiantes por Gestión"""
        # Si no se especifica gestión, usar año actual
        if not gestion:
            gestion = str(datetime.now().year)
        
        estudiantes = EstudianteRepository.obtener_por_gestion(
            db, gestion, nivel, id_curso, skip, limit
        )
        total = EstudianteRepository.contar_por_gestion(db, gestion, nivel, id_curso)
        
        estudiantes_dto = [EstudianteService._convertir_a_dto(est) for est in estudiantes]
        
        return EstudianteListDTO(
            total=total,
            estudiantes=estudiantes_dto
        )
    
    # ============= Importación/Exportación Excel =============
    
    @staticmethod
    def exportar_todos_excel(db: Session) -> BytesIO:
        """RF-EST-009: Exportar Todos los Estudiantes a Excel"""
        estudiantes = EstudianteRepository.obtener_todos_sin_paginacion(db)
        
        if not estudiantes:
            raise HTTPException(
                status_code=404,
                detail="No hay estudiantes para exportar"
            )
        
        # Crear libro de trabajo
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estudiantes"
        
        # Definir encabezados
        headers = [
            "ID", "CI", "Nombres", "Apellido Paterno", "Apellido Materno",
            "Fecha Nacimiento", "Dirección", "Estado",
            "Nombre Padre", "Apellido Paterno Padre", "Apellido Materno Padre", "Teléfono Padre",
            "Nombre Madre", "Apellido Paterno Madre", "Apellido Materno Madre", "Teléfono Madre"
        ]
        
        # Escribir encabezados con formato
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Escribir datos
        for row_num, est in enumerate(estudiantes, 2):
            ws.cell(row=row_num, column=1, value=est.id_estudiante)
            ws.cell(row=row_num, column=2, value=est.ci)
            ws.cell(row=row_num, column=3, value=est.nombres)
            ws.cell(row=row_num, column=4, value=est.apellido_paterno)
            ws.cell(row=row_num, column=5, value=est.apellido_materno)
            ws.cell(row=row_num, column=6, value=est.fecha_nacimiento.strftime("%Y-%m-%d") if est.fecha_nacimiento else "")
            ws.cell(row=row_num, column=7, value=est.direccion)
            ws.cell(row=row_num, column=8, value=est.estado)
            ws.cell(row=row_num, column=9, value=est.nombre_padre)
            ws.cell(row=row_num, column=10, value=est.apellido_paterno_padre)
            ws.cell(row=row_num, column=11, value=est.apellido_materno_padre)
            ws.cell(row=row_num, column=12, value=est.telefono_padre)
            ws.cell(row=row_num, column=13, value=est.nombre_madre)
            ws.cell(row=row_num, column=14, value=est.apellido_paterno_madre)
            ws.cell(row=row_num, column=15, value=est.apellido_materno_madre)
            ws.cell(row=row_num, column=16, value=est.telefono_madre)
        
        # Ajustar ancho de columnas
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def exportar_estudiante_excel(db: Session, id_estudiante: int) -> BytesIO:
        """RF-EST-010: Exportar Estudiante Individual a Excel"""
        estudiante = EstudianteRepository.obtener_por_id(db, id_estudiante)
        
        if not estudiante:
            raise HTTPException(
                status_code=404,
                detail=f"Estudiante con ID {id_estudiante} no encontrado"
            )
        
        # Crear libro de trabajo
        wb = openpyxl.Workbook()
        
        # === Hoja 1: Información del Estudiante ===
        ws_est = wb.active
        ws_est.title = "Estudiante"
        
        # Datos del estudiante
        datos = [
            ["Campo", "Valor"],
            ["ID", estudiante.id_estudiante],
            ["CI", estudiante.ci or ""],
            ["Nombres", estudiante.nombres],
            ["Apellido Paterno", estudiante.apellido_paterno],
            ["Apellido Materno", estudiante.apellido_materno or ""],
            ["Fecha Nacimiento", estudiante.fecha_nacimiento.strftime("%Y-%m-%d") if estudiante.fecha_nacimiento else ""],
            ["Dirección", estudiante.direccion or ""],
            ["Estado", estudiante.estado],
            ["", ""],
            ["Nombre Padre", estudiante.nombre_padre or ""],
            ["Apellido Paterno Padre", estudiante.apellido_paterno_padre or ""],
            ["Apellido Materno Padre", estudiante.apellido_materno_padre or ""],
            ["Teléfono Padre", estudiante.telefono_padre or ""],
            ["", ""],
            ["Nombre Madre", estudiante.nombre_madre or ""],
            ["Apellido Paterno Madre", estudiante.apellido_paterno_madre or ""],
            ["Apellido Materno Madre", estudiante.apellido_materno_madre or ""],
            ["Teléfono Madre", estudiante.telefono_madre or ""],
        ]
        
        for row_num, (campo, valor) in enumerate(datos, 1):
            ws_est.cell(row=row_num, column=1, value=campo).font = Font(bold=True)
            ws_est.cell(row=row_num, column=2, value=valor)
        
        ws_est.column_dimensions['A'].width = 25
        ws_est.column_dimensions['B'].width = 40
        
        # === Hoja 2: Cursos ===
        ws_cursos = wb.create_sheet(title="Cursos")
        headers_cursos = ["ID Curso", "Nombre Curso", "Nivel", "Gestión"]
        
        for col_num, header in enumerate(headers_cursos, 1):
            cell = ws_cursos.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for row_num, curso in enumerate(estudiante.cursos, 2):
            ws_cursos.cell(row=row_num, column=1, value=curso.id_curso)
            ws_cursos.cell(row=row_num, column=2, value=curso.nombre_curso)
            ws_cursos.cell(row=row_num, column=3, value=curso.nivel)
            ws_cursos.cell(row=row_num, column=4, value=curso.gestion)
        
        for col_num in range(1, len(headers_cursos) + 1):
            ws_cursos.column_dimensions[get_column_letter(col_num)].width = 20
        
        # Guardar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    async def importar_estudiantes_excel(db: Session, file: UploadFile) -> ImportResultDTO:
        """RF-EST-011: Importar Estudiantes desde Excel"""
        # Validar archivo
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=400,
                detail="El archivo debe ser Excel (.xlsx o .xls)"
            )
        
        # Validar tamaño (50 MB)
        contents = await file.read()
        if len(contents) > 50 * 1024 * 1024:
            raise HTTPException(
                status_code=400,
                detail="El archivo no debe superar 50 MB"
            )
        
        try:
            # Leer Excel con pandas
            df = pd.read_excel(BytesIO(contents))
            
            # Validar número de filas
            if len(df) > 10000:
                raise HTTPException(
                    status_code=400,
                    detail="El archivo no debe tener más de 10,000 filas"
                )
            
            # Validar columnas requeridas
            columnas_requeridas = ['CI', 'Nombres', 'Apellido Paterno', 'Apellido Materno']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                raise HTTPException(
                    status_code=400,
                    detail=f"Faltan columnas requeridas: {', '.join(columnas_faltantes)}"
                )
            
            # Procesar estudiantes
            creados = 0
            actualizados = 0
            errores = 0
            errores_detalle = []
            
            for idx, row in df.iterrows():
                try:
                    # Validar datos requeridos
                    if pd.isna(row['CI']) or not str(row['CI']).strip():
                        errores += 1
                        errores_detalle.append(f"Fila {idx + 2}: CI no proporcionado o vacío")
                        continue
                    
                    if pd.isna(row['Nombres']) or not str(row['Nombres']).strip():
                        errores += 1
                        errores_detalle.append(f"Fila {idx + 2}: Nombres no proporcionado")
                        continue
                    
                    if pd.isna(row['Apellido Paterno']) or not str(row['Apellido Paterno']).strip():
                        errores += 1
                        errores_detalle.append(f"Fila {idx + 2}: Apellido Paterno no proporcionado")
                        continue
                    
                    # Preparar datos
                    datos_estudiante = {
                        'ci': str(row['CI']).strip(),
                        'nombres': str(row['Nombres']).strip(),
                        'apellido_paterno': str(row['Apellido Paterno']).strip(),
                        'apellido_materno': str(row['Apellido Materno']).strip() if not pd.isna(row['Apellido Materno']) else None,
                    }
                    
                    # Campos opcionales
                    campos_opcionales = {
                        'Fecha Nacimiento': 'fecha_nacimiento',
                        'Dirección': 'direccion',
                        'Estado': 'estado',
                        'Nombre Padre': 'nombre_padre',
                        'Apellido Paterno Padre': 'apellido_paterno_padre',
                        'Apellido Materno Padre': 'apellido_materno_padre',
                        'Teléfono Padre': 'telefono_padre',
                        'Nombre Madre': 'nombre_madre',
                        'Apellido Paterno Madre': 'apellido_paterno_madre',
                        'Apellido Materno Madre': 'apellido_materno_madre',
                        'Teléfono Madre': 'telefono_madre',
                    }
                    
                    for col_excel, campo_db in campos_opcionales.items():
                        if col_excel in df.columns and not pd.isna(row[col_excel]):
                            valor = row[col_excel]
                            
                            # Convertir fecha
                            if campo_db == 'fecha_nacimiento' and isinstance(valor, pd.Timestamp):
                                valor = valor.date()
                            elif campo_db == 'fecha_nacimiento' and isinstance(valor, str):
                                try:
                                    valor = datetime.strptime(valor, "%Y-%m-%d").date()
                                except:
                                    valor = None
                            
                            datos_estudiante[campo_db] = valor
                    
                    # Crear o actualizar
                    _, fue_creado = EstudianteRepository.crear_o_actualizar_por_ci(
                        db, datos_estudiante
                    )
                    
                    if fue_creado:
                        creados += 1
                    else:
                        actualizados += 1
                
                except Exception as e:
                    errores += 1
                    errores_detalle.append(f"Fila {idx + 2}: {str(e)}")
            
            total_procesados = creados + actualizados + errores
            
            return ImportResultDTO(
                total_procesados=total_procesados,
                creados=creados,
                actualizados=actualizados,
                errores=errores,
                errores_detalle=errores_detalle[:20],  # Limitar a 20 errores
                mensaje=f"Importación completada: {creados} creados, {actualizados} actualizados, {errores} errores"
            )
        
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Error al procesar el archivo: {str(e)}"
            )
    
    @staticmethod
    def descargar_plantilla_excel() -> BytesIO:
        """RF-EST-012: Descargar Plantilla de Importación"""
        wb = openpyxl.Workbook()
        
        # === Hoja 1: Estudiantes (con ejemplo) ===
        ws = wb.active
        ws.title = "Estudiantes"
        
        headers = [
            "CI", "Nombres", "Apellido Paterno", "Apellido Materno",
            "Fecha Nacimiento", "Dirección", "Estado",
            "Nombre Padre", "Apellido Paterno Padre", "Apellido Materno Padre", "Teléfono Padre",
            "Nombre Madre", "Apellido Paterno Madre", "Apellido Materno Madre", "Teléfono Madre"
        ]
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Datos de ejemplo
        ejemplo = [
            "12345678", "Juan Carlos", "Pérez", "González",
            "2010-05-15", "Av. Principal #123", "Activo",
            "Carlos", "Pérez", "López", "70123456",
            "María", "González", "Mamani", "70654321"
        ]
        
        for col_num, valor in enumerate(ejemplo, 1):
            ws.cell(row=2, column=col_num, value=valor)
        
        # Ajustar columnas
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
        
        # === Hoja 2: Instrucciones ===
        ws_inst = wb.create_sheet(title="Instrucciones")
        
        instrucciones = [
            ["INSTRUCCIONES PARA IMPORTAR ESTUDIANTES", ""],
            ["", ""],
            ["Columnas OBLIGATORIAS:", ""],
            ["- CI", "Cédula de identidad del estudiante (debe ser única)"],
            ["- Nombres", "Nombres completos del estudiante"],
            ["- Apellido Paterno", "Apellido paterno del estudiante"],
            ["- Apellido Materno", "Apellido materno del estudiante"],
            ["", ""],
            ["Columnas OPCIONALES:", ""],
            ["- Fecha Nacimiento", "Formato: YYYY-MM-DD (ej: 2010-05-15)"],
            ["- Dirección", "Dirección completa del estudiante"],
            ["- Estado", "Valores válidos: Activo, Retirado, Abandono"],
            ["- Información del Padre", "Nombre, apellidos y teléfono"],
            ["- Información de la Madre", "Nombre, apellidos y teléfono"],
            ["", ""],
            ["NOTAS IMPORTANTES:", ""],
            ["1.", "Si el CI ya existe, se actualizará el estudiante"],
            ["2.", "Si el CI no existe, se creará un nuevo estudiante"],
            ["3.", "El archivo no debe superar 10,000 filas"],
            ["4.", "El tamaño máximo del archivo es 50 MB"],
            ["5.", "Use la hoja 'Estudiantes' como referencia"],
        ]
        
        for row_num, (col1, col2) in enumerate(instrucciones, 1):
            cell1 = ws_inst.cell(row=row_num, column=1, value=col1)
            ws_inst.cell(row=row_num, column=2, value=col2)
            
            if row_num == 1:
                cell1.font = Font(bold=True, size=14, color="4472C4")
            elif col1 in ["Columnas OBLIGATORIAS:", "Columnas OPCIONALES:", "NOTAS IMPORTANTES:"]:
                cell1.font = Font(bold=True, size=12)
        
        ws_inst.column_dimensions['A'].width = 30
        ws_inst.column_dimensions['B'].width = 60
        
        # Guardar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    # ============= Utilidades =============
    
    @staticmethod
    def _convertir_a_dto(estudiante) -> EstudianteResponseDTO:
        """Convertir modelo a DTO"""
        cursos_dto = [
            CursoBasicoDTO(
                id_curso=curso.id_curso,
                nombre_curso=curso.nombre_curso,
                nivel=curso.nivel,
                gestion=curso.gestion
            )
            for curso in estudiante.cursos
        ]
        
        return EstudianteResponseDTO(
            id_estudiante=estudiante.id_estudiante,
            ci=estudiante.ci,
            nombres=estudiante.nombres,
            apellido_paterno=estudiante.apellido_paterno,
            apellido_materno=estudiante.apellido_materno,
            fecha_nacimiento=estudiante.fecha_nacimiento,
            direccion=estudiante.direccion,
            estado=estudiante.estado,
            nombre_padre=estudiante.nombre_padre,
            apellido_paterno_padre=estudiante.apellido_paterno_padre,
            apellido_materno_padre=estudiante.apellido_materno_padre,
            telefono_padre=estudiante.telefono_padre,
            nombre_madre=estudiante.nombre_madre,
            apellido_paterno_madre=estudiante.apellido_paterno_madre,
            apellido_materno_madre=estudiante.apellido_materno_madre,
            telefono_madre=estudiante.telefono_madre,
            cursos=cursos_dto
        )
